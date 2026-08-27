#!/usr/bin/env python3
"""Refresh social/calendar.json from the root xlsx. Does not delete anything."""
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "ORION_Publishing_Calendar.xlsx"
OUT = Path(__file__).resolve().parents[1] / "calendar.json"

KEYS = {
    "Date": "date",
    "Weekday": "weekday",
    "Format": "format",
    "Title / Mutator": "title",
    "Hook / Concept": "hook",
    "Sunday Double": "sundayDouble",
    "Asset filename": "asset",
    "Asset status": "assetStatus",
    "Post status": "postStatus",
    "Instagram caption": "igCaption",
    "Instagram tags": "igTags",
    "TikTok caption": "ttCaption",
    "TikTok note": "ttNote",
    "YouTube title": "ytTitle",
    "YouTube description": "ytDescription",
}


def main():
    ws = openpyxl.load_workbook(XLSX, data_only=True).active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    posts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw = dict(zip(headers, row))
        if not raw.get("Date"):
            continue
        d = raw["Date"]
        date = d.date().isoformat() if hasattr(d, "date") else str(d)[:10]
        rec = {}
        for h, k in KEYS.items():
            v = raw.get(h)
            if k == "date":
                rec[k] = date
            elif k == "sundayDouble":
                rec[k] = bool(v) and str(v).strip().lower() in ("yes", "true", "1")
            else:
                rec[k] = None if v is None else str(v)
        posts.append(rec)
    OUT.write_text(
        json.dumps({"source": XLSX.name, "posts": posts}, indent=2, ensure_ascii=False) + "\n"
    )
    print(f"{len(posts)} rows -> {OUT}")


if __name__ == "__main__":
    main()
