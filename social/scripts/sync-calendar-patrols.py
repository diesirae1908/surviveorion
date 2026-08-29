#!/usr/bin/env python3
"""Rewrite TODAY'S PATROL rows from the live mutator pool. Does not delete rows."""
import json
import subprocess
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
XLSX = ROOT / "ORION_Publishing_Calendar.xlsx"
EXPORT = Path(__file__).resolve().parent / "export-calendar.py"
CUTOFF = "2026-08-29"


def slug(names: list[str]) -> str:
    return "".join(n.lower().replace(" ", "") for n in names)


def weekday_short(date: str) -> str:
    from datetime import date as Date

    y, m, d = map(int, date.split("-"))
    return Date(y, m, d).strftime("%a")


def patrol_cells(pick: dict) -> dict:
    names = pick["names"]
    title_names = " + ".join(names)
    hook = " ".join(pick["briefings"])
    subline = " ".join(pick["sublines"])
    sunday = len(names) > 1
    title = f"PATROL: {title_names}" if sunday else title_names
    fmt = "TODAY'S PATROL · SUNDAY DOUBLE" if sunday else "TODAY'S PATROL"
    mmdd = pick["date"][5:7] + pick["date"][8:10]
    asset = f"{mmdd}_patrol_{slug(names)}_916.mov"
    ig = (
        f"Today every pilot flies {title_names}: {hook} "
        "Three attempts. Free, in your browser."
    )
    tt = ig.lower()
    tags = "#dailychallenge #indiegame #browsergame #arcadegame"
    yt_title = f"{title_names} day | daily dodge game (ORION Day {pick['dayNumber']})"
    yt_desc = (
        f"Today's shared seed: {title_names}. {subline} "
        "Every pilot flies the same run. Three attempts, free in your browser.\n\n"
        "ORION is a daily dodge game. Same swarm as every other pilot today.\n\n"
        "Play now: https://surviveorion.com\n\n"
        f"{tags} #Shorts"
    )
    return {
        "Date": pick["date"],
        "Weekday": weekday_short(pick["date"]),
        "Format": fmt,
        "Title / Mutator": title,
        "Hook / Concept": hook,
        "Sunday Double": "Yes" if sunday else None,
        "Asset filename": asset,
        "Asset status": "Not linked yet",
        "Post status": "Draft",
        "Instagram caption": ig,
        "Instagram tags": tags,
        "TikTok caption": tt,
        "TikTok note": None,
        "YouTube title": yt_title,
        "YouTube description": yt_desc,
    }


def dump_schedule(from_date: str, days: int) -> list[dict]:
    raw = subprocess.check_output(
        ["npx", "tsx", "scripts/dump-patrol-schedule.ts", from_date, str(days)],
        cwd=ROOT,
        text=True,
    )
    return json.loads(raw)


def apply_flood_posted_state(ws, headers: list) -> None:
    """calendar.json already shipped THE FLOOD; keep that approval on the xlsx."""
    for row in ws.iter_rows(min_row=2):
        rec = {headers[i]: row[i] for i in range(len(headers))}
        date = rec["Date"].value
        date = date.date().isoformat() if hasattr(date, "date") else str(date)[:10]
        fmt = str(rec["Format"].value or "")
        if date == "2026-08-28" and fmt.startswith("TODAY'S PATROL"):
            rec["Asset filename"].value = "0827_theflood_916.mov"
            rec["Asset status"].value = "Linked"
            rec["Post status"].value = "Approved"


def main() -> int:
    days = int(sys.argv[1]) if len(sys.argv) > 1 else 14
    picks = {p["date"]: p for p in dump_schedule(CUTOFF, days)}
    wb = openpyxl.load_workbook(XLSX)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    updated = []
    for row in ws.iter_rows(min_row=2):
        rec = {headers[i]: row[i] for i in range(len(headers))}
        date = rec["Date"].value
        date = date.date().isoformat() if hasattr(date, "date") else str(date)[:10]
        fmt = str(rec["Format"].value or "")
        if date < CUTOFF or not fmt.startswith("TODAY'S PATROL"):
            continue
        pick = picks.get(date)
        if not pick:
            continue
        cells = patrol_cells(pick)
        for h, v in cells.items():
            rec[h].value = v
        updated.append(f"{date} {cells['Title / Mutator']}")
    apply_flood_posted_state(ws, headers)
    wb.save(XLSX)
    subprocess.check_call([sys.executable, str(EXPORT)])
    print("updated:")
    for line in updated:
        print(f"  {line}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
